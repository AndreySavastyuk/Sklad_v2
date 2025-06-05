from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse
from typing import List, Optional
import sqlite3
from pathlib import Path
from openpyxl import load_workbook

from .database import get_connection, init_db
from .schemas import Task, TaskCreate, TaskUpdate

app = FastAPI(title="Warehouse Task Server")

STATIC_DIR = Path(__file__).parent / "public"


@app.on_event("startup")
def startup():
    init_db()


@app.get("/", response_class=HTMLResponse)
def index():
    html_file = STATIC_DIR / "index.html"
    if html_file.exists():
        return html_file.read_text(encoding="utf-8")
    return "<h1>Warehouse Task Server</h1>"


@app.post("/tasks", response_model=Task)
def create_task(task: TaskCreate):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(
            """INSERT INTO tasks (task_number, comment, assembly_count, created_by)
               VALUES (?, ?, ?, ?)""",
            (task.task_number, task.comment, task.assembly_count, task.created_by),
        )
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        raise HTTPException(status_code=409, detail="Task number already exists")

    task_id = cursor.lastrowid
    row = cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    conn.close()
    return Task(**row)


@app.get("/tasks", response_model=List[Task])
def list_tasks(status: Optional[str] = None):
    conn = get_connection()
    cursor = conn.cursor()
    if status:
        rows = cursor.execute(
            "SELECT * FROM tasks WHERE status = ? ORDER BY created_at DESC", (status,)
        ).fetchall()
    else:
        rows = cursor.execute("SELECT * FROM tasks ORDER BY created_at DESC").fetchall()
    conn.close()
    return [Task(**row) for row in rows]


@app.get("/tasks/{task_id}", response_model=Task)
def get_task(task_id: int):
    conn = get_connection()
    row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    conn.close()
    if row:
        return Task(**row)
    raise HTTPException(status_code=404, detail="Task not found")


@app.put("/tasks/{task_id}", response_model=Task)
def update_task(task_id: int, task: TaskUpdate):
    conn = get_connection()
    cursor = conn.cursor()
    fields = []
    values = []
    if task.comment is not None:
        fields.append("comment = ?")
        values.append(task.comment)
    if task.assembly_count is not None:
        fields.append("assembly_count = ?")
        values.append(task.assembly_count)
    if task.status is not None:
        fields.append("status = ?")
        values.append(task.status)
    if not fields:
        raise HTTPException(status_code=400, detail="No fields to update")
    fields.append("updated_at = CURRENT_TIMESTAMP")
    sql = f"UPDATE tasks SET {', '.join(fields)} WHERE id = ?"
    values.append(task_id)
    cursor.execute(sql, values)
    conn.commit()
    row = cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    conn.close()
    if row:
        return Task(**row)
    raise HTTPException(status_code=404, detail="Task not found")


@app.put("/tasks/{task_id}/status", response_model=Task)
def update_status(task_id: int, status: str):
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE tasks SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (status, task_id),
    )
    conn.commit()
    row = cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    conn.close()
    if row:
        return Task(**row)
    raise HTTPException(status_code=404, detail="Task not found")


@app.post("/tasks/upload")
def upload_excel(file: UploadFile = File(...)):
    wb = load_workbook(file.file)
    sheet = wb.active
    conn = get_connection()
    cursor = conn.cursor()
    created_ids = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        task_number = str(row[0])
        comment = str(row[1]) if row[1] else None
        cursor.execute(
            "INSERT INTO tasks (task_number, comment) VALUES (?, ?)",
            (task_number, comment),
        )
        created_ids.append(cursor.lastrowid)
    conn.commit()
    conn.close()
    return {"created": created_ids}


@app.post("/tasks/parse")
def parse_excel(file: UploadFile = File(...)):
    """Parse Excel file and return raw rows for client-side editing."""
    wb = load_workbook(file.file)
    sheet = wb.active
    headers = [cell.value or f"col{i+1}" for i, cell in enumerate(sheet[1])]
    items = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: (str(val) if val is not None else "") for i, val in enumerate(row)}
        items.append(item)
    return {"headers": headers, "items": items}
